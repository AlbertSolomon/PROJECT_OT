from openpyxl import Workbook, load_workbook
from ot import OT
from lang import path, overtime_month, delisting, settings_path, saved_file_location, job_titles
import json
from datetime import date
import re

#saved_file_location()
workbook = load_workbook(path)
workbook.active

#print(type(workbook.sheetnames))
#print(workbook.sheetnames)

def command_prompter() -> list:
    """This function only exists for testing purposes """
    ot_details: list= []
    month = input("enter month for recording/retrieving ovetime details:", )
    sheet = input("enter sheet name:")
    default_sheet = delisting(workbook.sheetnames)
    ot_details.append(month)

    ot_details.append(sheet if sheet !='' else default_sheet)
    OT(workbook, sheet if sheet !='' else default_sheet).delete_empty_rows() #? search_job_description(job_titles.keys())
    print(date.today())
    return ot_details

command_prompter()
# workbook.save('./final/june.xlsx')
workbook.save(f"./final/{ overtime_month }.xlsx")