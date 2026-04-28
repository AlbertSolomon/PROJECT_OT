from lang import get_ot_data, path, t_path, settings_path, rules_path, path_blank, overtime_month, rules_settings
from ot import OT
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.formula import ArrayFormula # example implementation at https://openpyxl.readthedocs.io/en/stable/simple_formulae.html#id2

# workbook = load_workbook(path_blank) #path
workbook = load_workbook(path_blank)
workbook.active

def sync_json_DATA():
    ot_data = get_ot_data()
    print(rules_settings())
    

sync_json_DATA()

#def sync_json_data():
#    # might use a helper function but who knows
#    ot_data = get_ot_data()
#    data_list:list = []
#    record_dict: dict = {}
#    counter: int = 0
#    # print(ot_data)
#    # print(type(ot_data))
#    for key in ot_data:
#        try:
#            # print(ot_data[key])
#            temp_workbook = OT(workbook, key)
#            start_point = str(int(temp_workbook.insert_point_coordinate().lstrip('A')) - 1)
#            print("loop:",start_point)
#            print(type(temp_workbook))
#            print(f"{ key }: is", temp_workbook.finder())
#
#            if temp_workbook.finder():
#                print(len(ot_data[key]))
#                print(ot_data[key])
#                data_list = ot_data[key]
#                for _, indx in enumerate(data_list):
#                    #print(data_list[_]
#                    record_dict = data_list[_]
#                    #print(temp_workbook.insert_point_coordinate())
#                    #print(start_point)
#                    # print(f"index ::: { indx }",record_dict.get("date"))
#
#                # this is from an alternate branch 
#                # we are first checking the records the sheet haas 
#                # we are inserting the rows according to the amount records in a sheet from a specific a specific location.
#
#            #workbook.save(f"./temp/{ overtime_month }.xlsx")
#        except KeyError:
#            print(f"{key} does not exist in the workbook")
#            # create a workbook
#            # count number of CELLS IN ROW A before TOTAL, CHECK CELL A4 FOR EMPTYNESS AND START THERE (recording)
#            # add records to the workbook
#            # save to a temporary file 
#
#    # sheetname, DATE, DAY, PURPOSE (depends with the posistion ),START, FINISH, TOTAL HOURS, RATE(always 1.5)
#    # if self.finder(key):
#    #    sheet = self.workbook[key]
#
## sync_json_data()
